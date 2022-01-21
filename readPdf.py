import flask
import os.path
import PyPDF2
import json as json_dict
import json
import jsonschema
import pandas as pd
import re
import camelot.io as camelot
import tabula
import codecs
import random
import base64
import pikepdf
import lxml.etree as ET
from pandas.io.json import json_normalize
from lxml import etree, objectify
from lxml.etree import XMLSyntaxError
from jsonschema import validate
from collections import OrderedDict
from bs4 import BeautifulSoup
from flask import request, jsonify
from datetime import datetime
from imap_tools import MailBox, AND
from email.message import EmailMessage
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
from datetime import date
import shutil
import pathlib
import win32com.client
import openpyxl
import xmltodict, json
import pdfkit
import ctypes
from ctypes.util import find_library

app = flask.Flask(__name__)
app.config["DEBUG"] = True

@app.route('/api', methods=['POST'])
def home():
    #pd.show_versions()
    #if request.method == 'POST':
    # json = request.json
    # ruta = json['file']
    # resultado = ''
    # pdfFileObj = open(ruta,'rb')
    # pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
    # print("Cantidad de páginas: ", pdfReader.numPages)
    # print("Encriptado: ", pdfReader.isEncrypted)
    # info = pdfReader.documentInfo
    # print("info: ", info)
    # for x in range(0, pdfReader.numPages):
        # pageObj = pdfReader.getPage(x)
        # resultado += pageObj.extractText()
    # pdfFileObj.close()
    return find_library("".join(("gsdll", str(ctypes.sizeof(ctypes.c_voidp) * 8), ".dll")))
    #return request.args['file']

@app.route('/DataFromTable', methods=['POST'])
def getDataFromTable():
    datasets = []
    json = request.json
    html_data = json['html']
    esquema = json['esquema']
    tabla = table_json(html_data, esquema)
    print("esquema")    
    print(esquema)

    if  not (tabla is None):
       headings = [th.get_text() for th in tabla.find("tr").find_all("span")]
       print(headings)
       
       for row in tabla.find_all("tr")[1:]:
           dataset = dict(zip(headings, (td.get_text() for td in row.find_all("td"))))
           datasets.append(dataset)

    print(datasets)
    return str(datasets)
@app.route('/json_to_html', methods=['POST'])
def json_to_html():
    entrada = request.json
    v_json = entrada['v_json']
    with open(v_json, encoding="utf8") as json_file:
        data = json.load(json_file)

    df_generales = pd.json_normalize(data).T.loc[['data.nroreq','data.propietario','data.propietariocorreo','data.fechainicio','data.fechafin',
                              'data.estado','data.notas','data.comentarios','data.direccion']]
    df_titulo = pd.json_normalize(data).T.loc[['data.nomorgcompradora']]
    df_generales.index = ['Nro Requerimiento','Contacto','Correo','Fecha Inicio','Fecha Fin','Estado','Notas','Comentarios','Dirección']
    df_generales.columns = ['Dato']
    df_atributos = pd.json_normalize(data, record_path =['data','atributos'])
    df_atributos = df_atributos.loc[df_atributos.nombreatributo.isin(['Moneda','Validez de la Oferta']),['nombreatributo','valor']]
    df_atributos = df_atributos.set_index('nombreatributo')
    df_atributos.columns = ['Dato']
    df_generales = df_generales.append(df_atributos)
    df_producto = pd.json_normalize(data, record_path =['data','productos'])
    df_producto = df_producto[['idproducto','codigoproducto','nombreproducto','descripcionproducto']]
    df_producto.columns = ['RFQ','Codigo Producto','Nombre Producto','Descripción Producto']
    df_producto_2 = pd.json_normalize(data, record_path =['data','productos','atributoxproducto'])[['nombreatributo','valorenviado','idproductoxrfq']]
    df_producto_2 = df_producto_2.merge(df_producto, left_on='idproductoxrfq', right_on='RFQ').pivot(index="nombreatributo", columns="idproductoxrfq", values="valorenviado")
    df_producto_2 = df_producto_2.T
    df_producto_2 = df_producto_2[['Cantidad','Comentarios','Material','Numero de Parte','Observaciones','PosiciÃ³n','Unidad de Medida']]
    df_producto_2.columns = ['Cantidad','Comentarios','Material','Numero de Parte','Observaciones','Posicion','Unidad de Medida']
    df_producto = df_producto.reset_index(drop=True)
    df_producto_2 = df_producto_2.reset_index(drop=True)
    df_producto = df_producto.join(df_producto_2)
    df_producto = df_producto.reset_index(drop=True)
    df_producto = df_producto[['Posicion','Codigo Producto','Nombre Producto','Descripción Producto',
                    'Cantidad','Unidad de Medida','Numero de Parte','Material','Comentarios','Observaciones']]
    html = df_generales.to_html(justify = 'left')
    html2 = df_producto.to_html(index = False, justify = 'left', classes='table table-striped')
    # html2 = html2.replace('dataframe ','')
    #write html to file
    text_file = open(v_html, "w")
    html3 = ("<h1><center><b>" + str(df_titulo.values[0]) + "</b></center></h1><p style='color:#0000ff';><b>Detalles Generales</b></p>" + html+ "<br><p style='color:#0000ff';><b>Detalles de Productos</b></p>" + html2)
    #text_file.write(html3)
    #text_file.close()
    v_pdf = os.path.splitext(v_json)[0] + '.pdf'
    pdfkit.from_string(html3, v_pdf)    
    return v_pdf

@app.route('/HtmlToPDF', methods=['POST'])
def HtmlToPdf():
    entrada = request.json
    html = entrada['html']
	
    pdf = entrada['pdf']
    
    soup = BeautifulSoup(html, 'html.parser')
    img = soup.find_all('img')
    for t in img:
        if t:  # safety feature
           t.extract()
    a = soup.find_all('a')
    for t in a:
        if t:  # safety feature
           t.extract()
    html = str(soup)
    html_content = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
</head>
<body>
    """+html+"""
<body>
</html>
"""
    #html = html.encode("utf-8").decode("latin-1")
    config = pdfkit.configuration(wkhtmltopdf="C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe")
    pdfkit.from_string(html_content, pdf, configuration=config) 
    return pdf
@app.route('/TableFromPDF', methods=['POST'])
def getTableFromPDF():
    json = request.json
    pdf = json['pdf']
    concatenar = json['concatenar']
    #esquema = json['esquema']
    rnd = random.random()*100000000000
    nombre_pdf = str(rnd) + ".pdf"
    with open(nombre_pdf, 'wb') as theFile:
         theFile.write(base64.b64decode(pdf))
    with pikepdf.open(nombre_pdf, allow_overwriting_input=True) as pdf:
        pdf.save(nombre_pdf)
    tables = camelot.read_pdf(nombre_pdf,pages='1-end', line_scale=40, shift_text=[''])
    print(tables)
    i = 1
    lines = ''
    if concatenar == 1:
        for t in tables:
            soup = BeautifulSoup(t.df.to_html(), 'html.parser')
            table_rows = soup.find_all('th')
            for th in table_rows:
                if th:  # safety feature
                   th.extract()
            table_rows = soup.find_all('tr')
            j = 1
            for tr in table_rows:
                if i == 1:
                   lines += str(tr)
                else:
                   table_column = tr.find_all('td')
                   if len(table_column) > 1:
                      if table_column[0].text.strip():
                         if j > 1:
                            lines += str(tr)
                         j += 1
            i += 1
        lines = "<table>"+lines+"</table>"
    else:
        for t in tables:
            soup = BeautifulSoup(t.df.to_html(), 'html.parser')
            table_rows = soup.find_all('th')
            for th in table_rows:
                if th:  # safety feature
                   th.extract()
            lines += str(soup)
            i += 1
    
    soup = BeautifulSoup(lines, 'html.parser')
    
    return soup.prettify()

@app.route('/TableFromPDFTabula', methods=['POST'])
def getTableFromPDFTabula():
    json = request.json
    pdf = json['pdf']
    data = TableColumnPDFTabula(pdf)
    return data
def TableColumnPDFTabula(pdf):
    rnd = random.random()*100000000000
    nombre_pdf = str(rnd) + ".pdf"
    with open(nombre_pdf, 'wb') as theFile:
         theFile.write(base64.b64decode(pdf))
    tables = camelot.read_pdf(nombre_pdf,pages='1-end' , flavor='stream', edge_tol=500)#, line_scale=40, shift_text=[''])
    #print(len(tables))
    i = 0
    lines = ''
    data = {}
    for t in tables:
        t.df.loc[(t.df[2] == "") & (t.df[3].str.len() > 15), 2] = t.df[2].str.cat(t.df[3], sep =" ")
        t.df.loc[(t.df[0] == 'Nº Ref. Proveedor'), 0] = ""
        t.df.loc[(t.df[1] == "") & (t.df[0].str.isnumeric() == False) & (t.df[0] != 'Nº Ref. Proveedor'), 1] = t.df[0]
        t.df.loc[(t.df[3].str.len() > 18), 3] = ""
        t.df.loc[(t.df[1] != "") & (t.df[1].str.isnumeric() == False), 1] = t.df[1].str.cat(t.df[3], sep ="@")
        t.df.loc[(t.df[2].str.isnumeric() == True), 1] = t.df[2]
        t.df.loc[(t.df[2].str.isnumeric() == True), 2] = ""
        t.df.loc[(t.df[2] == ""), 2] = t.df[3]        
        t.df.loc[(t.df[0] == '.'), 0] = ""
        t.df.loc[(t.df[1] == '.|'), 1] = ""
        t.df.loc[(t.df[1] == '.,'), 1] = ""
        t.df.drop(t.df[(t.df[0] == "") & (t.df[1] == "") & (t.df[2] == "") & (t.df[3] == "")].index, inplace=True)
        t.df.loc[(t.df.apply(lambda x: x[2] in x[1], axis=1)), 2] = ""
        t.df.loc[(t.df.apply(lambda x: x[3] in x[2], axis=1)), 3] = ""
        t.df.loc[(t.df[0] != "") & (t.df[0].str.isnumeric() == True), 2] = t.df[0].str.cat(t.df[2], sep ="@")
        soup = BeautifulSoup(t.df.to_html(), 'html.parser')
        table_rows = soup.find_all('th')
        for th in table_rows:
            if th:  # safety feature
               th.extract()
        table_rows = soup.find_all('tr')
        salir = 0
        
        for tr in table_rows:
            for td in tr.find_all('td'):
                if td.get_text() != "Item":
                    tr.extract()
                else: 
                    salir = 1
                    break
            if salir == 1:
               break
        table_rows = soup.find_all('tr')
        salir = 0
        aux_salto = 0
        for tr in table_rows:
            for td in tr.find_all('td'):
               #print( td.get_text())
               #print("--------------------------------------")
               if "Fabricante" in td.get_text(): 
                  print("Encontrado")
                  salir = 1
                  break
            if salir == 1:
               break
            aux_salto += 1 
        print("salto: " + str(aux_salto))
        lines += str(soup)
        total_columnas = 0
        for row in soup.find_all('tr'):
           columns = row.find_all('td')
           if len(columns) > 0:
              total_columnas = len(columns)
              break
        for x in range(0, total_columnas):
           #print(str(x))
           salto = 0
           data['columna' + str(i)] = []
           for row in soup.find_all('tr'):
               columns = row.find_all('td')
               if columns and salto > aux_salto:
                  c = columns[x]
                  if c.get_text():
                     if "Favor revisar nuestros términos y condiciones" in c.get_text():
                         break
                     else:
                         data['columna' + str(x)].append(c.get_text())
                  #print(data)
               salto += 1
           i += 1
    
    return data
@app.route('/ContentFromPDF', methods=['POST'])
def ContentFromPDF():
    json = request.json
    pdf = json['pdf']
    rnd = random.random()*100000000000
    nombre_pdf = str(rnd) + ".pdf"
    with open(nombre_pdf, 'wb') as theFile:
         theFile.write(base64.b64decode(pdf))
    tables = camelot.read_pdf(nombre_pdf,pages='1-end' , flavor='stream', edge_tol=5)#, line_scale=40, shift_text=[''])
    i = 0
    lines = ''
    data = {}
    data["Productos"] = []
    data["RFQNumber"] = []
    data["Fecha"] = []
    data["Comprador"] = []
    linea = ""
    
    for t in tables:
        soup = BeautifulSoup(t.df.to_html(), 'html.parser')
        table_rows = soup.find_all('th')
        for th in table_rows:
            if th:  # safety feature
               th.extract()
        table_rows = soup.find_all('tr')
        salir = 0
        texto = ""
        for tr in table_rows:
            for td in tr.find_all('td'):
                if td.get_text() != "Item":
                    texto = td.get_text().split("\\n")                    
                    if len(texto) == 2 and texto[0].isnumeric():
                        data["RFQNumber"] = texto[0]                        
                    if 'Buyer / Comprador:' in td.get_text():
                       comprador = td.get_text().replace("Buyer / Comprador:", "")
                       data["Comprador"] = comprador
                    if 'Quotation Deadline / Fecha de Cierre:' in td.get_text():
                       fecha_cierre = td.get_text().replace("Quotation Deadline / Fecha de Cierre:", "")
                       fecha_cierre = fecha_cierre.replace("\\n","")
                       
                       date_time_obj = fecha_cierre
                       data["Fecha"]= date_time_obj
                    tr.extract()
                else: 
                    salir = 1
                    break
            if salir == 1:
               break
        table_rows = soup.find_all('tr')
        salir = 0
        aux_salto = 0
        for tr in table_rows:
            for td in tr.find_all('td'):
               if "Fabricante" in td.get_text(): 
                  print("Encontrado")
                  salir = 1
                  break
               else:
                   tr.extract()
            if salir == 1:
               break
            aux_salto += 1 
        print("salto: " + str(aux_salto))
        for th in soup.find_all('thead'):
            th.extract()

        
        lines += str(soup)
        salir = 0
        valor_celda = ""
        for tr in soup.find_all('tr'):
            for td in tr.find_all('td'):  
                valor_celda = td.get_text() 
                break
            break
        print("valor_celda: " + valor_celda)
        
        if valor_celda:
            for row in soup.find_all('tr'):
                for td in row.find_all('td'):                
                    if td:
                        if td.get_text() and not 'Favor revisar nuestros términos' in td.get_text() and not 'Por favor revisar' in td.get_text() and not 'Please review our' in td.get_text() and not 'https://publicportal.fmi.com/cerro_verde' in td.get_text() and not 'Submitted by' in td.get_text():
                            contenido = td.get_text().split('\\n')
                            for texto in contenido:
                                if texto:
                                    linea += texto + "@"
    data["Productos"] = linea[0:len(linea) - 1]
    print(lines)
    return data
@app.route('/robotveritrade', methods=['POST'])
def robotveritrade():
    json = request.json
    marca = json['marca']
    year = json['anio']
    directorio = json['directorio']
    user = json['user']
    password = json['pass']
    default_directory = "C:\\temp"
    #directorio = "E:\\Documentos\\EFC\\VERITRADE"
    result = {}
    result['filename'] = ""
    result['error'] = ""
    chrome_options = Options()
    #chrome_options.add_argument("--headless")
    chrome_options.add_argument("--window-size=1920x1080")
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--verbose')
    chrome_options.add_experimental_option("prefs", {
            "download.default_directory": default_directory,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing_for_trusted_sources_enabled": False,
            "safebrowsing.enabled": False
    })
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--disable-software-rasterizer')
    
    driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options)
    enable_download_headless(driver, default_directory)
    driver.maximize_window()
    data = []
    url = 'https://www.veritradecorp.com/'
    driver.get(url)
    print("Ingresando a " + url)
    time.sleep(5)
    
    
    div_login = driver.find_element_by_id("li_login")
    div_login.click()
    print("Click en div login")

    Login_Rut = driver.find_element_by_id("txtCodUsuario")
    Login_Rut.click()
    #Login_Rut.send_keys("CREMUZGO@EFC.COM.PE")
    Login_Rut.send_keys(user)
    print("Se envia username")
    Login_Clave = driver.find_element_by_id("txtPassword")
    Login_Clave.click()
    #Login_Clave.send_keys("545CHR")
    Login_Clave.send_keys(password)
    print("Se envia password")
    Btn_Ingresar = driver.find_element_by_id("login_button")
    Btn_Ingresar.click()
    print("Click en botón login")
    time.sleep(10)
    
    existe_usuario = driver.find_elements_by_xpath("/html/body/div[2]/div/div/div[2]/b[1]")
    print("Se busca mensaje si existe otro usuario conectado.")
    print(len(existe_usuario))
    if (len(existe_usuario) == 0):
        try:
            btnNoShowMeAviso = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "btnNoShowMeAviso")))
            if (btnNoShowMeAviso):
                btnNoShowMeAviso = driver.find_element_by_id("btnNoShowMeAviso")
                btnNoShowMeAviso.click()
            else:
                print("no existe")
        except Exception as e:
            print(e)
            print("existe")
        
        try:
            print("NO existe otro usuario conectado.")
            pais = driver.find_elements_by_xpath("/html/body/div[2]/div[2]/div/div[1]/div[1]/select[2]/option[19]")
            if (len(pais) > 0):
                pais[0].click()
                print("Click en páis Perú")
                time.sleep(3)
            
            filtro = driver.find_element_by_id("txtDesComercialB")
            filtro.click()
            filtro.send_keys(marca)
            print("Se escribe marca en filtro")
            
            btnAgregarDesComercial = driver.find_element_by_id("btnAgregarDesComercial")
            btnAgregarDesComercial.click()
            print("Se agrega filtro")
            time.sleep(3)
            
            today = date.today()
            if not year:
               year = 5
            anio = today.year - int(year)
            mes = today.month - 1
            driver.execute_script("$('#cboDesde').datepicker('setDate', new Date("+str(anio)+","+str(mes)+",01));")
            print("Se setea fecha desde a "+str(year)+" años atrás")
            btnBuscar = driver.find_element_by_id("btnBuscar")
            btnBuscar.click()
            print("Click en buscar")
            time.sleep(30)
            popupSinResultado = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "message")))
            popupSinResultado = driver.find_element_by_id("message")
            
            if not popupSinResultado.get_attribute("innerHTML"):
                result['error'] = popupSinResultado.get_attribute("innerHTML")
            else:
                tabDetalleExcel = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "tabDetalleExcel")))
                tabDetalleExcel = driver.find_element_by_id("tabDetalleExcel")
                tabDetalleExcel.click()
                print("Click en tab detalle excel")
                time.sleep(20)
                
                downloadFileVerRegistro2 = driver.find_element_by_id("downloadFileVerRegistro2")
                downloadFileVerRegistro2.click()
                print("Click en descargar excel")
                time.sleep(70)
                
                fecha_archivo = today.strftime("%Y%m%d")
                filename = rename_last_downloaded_file(default_directory, default_directory, marca + fecha_archivo)
                print("Se obtiene archivo descargado: " + filename)
                file_extension = pathlib.Path(filename).suffix
                aux = 5
                if (file_extension == ".xls"):
                    aux = 0
                    print("Procesando XLS.")
                    read_file = pd.read_csv(filename, sep='\t', encoding= 'unicode_escape', error_bad_lines=False)
                    print("Archivo XLS leído.")
                    filename = os.path.join(default_directory, marca + fecha_archivo + '.xlsx')
                    print("filename xlsx: " + filename)
                    read_file.to_excel(filename, index = None, header=True)
                    
                    
                    wb = openpyxl.load_workbook(filename)
                    sheet = wb.worksheets[0]
                    sheet.insert_cols(4)
                    sheet.cell(row=1, column=4).value = "DUA REP"
                    sheet.insert_cols(6)
                    sheet.cell(row=1, column=6).value = "AÑO"
                    print("Se agregan columnas adicionales DUA REP y AÑO")
                    
                    for i in range(2, sheet.max_row+1):
                        sheet.cell(row=i, column=4).value = sheet.cell(row=i, column=5).value[0:6]
                        try:
                            sheet.cell(row=i, column=6).value = sheet.cell(row=i, column=7).value[0:4]
                        except:
                            print("")
                    print("Se aplican formulas para columnas agregadas.")
                else:
                    wb = openpyxl.load_workbook(filename)
                    sheet = wb.worksheets[0]
                    sheet.insert_cols(4)
                    sheet.cell(row=6, column=4).value = "DUA REP"
                    sheet.insert_cols(6)
                    sheet.cell(row=6, column=6).value = "AÑO"
                    print("Se agregan columnas adicionales DUA REP y AÑO")
                    
                    # iterate through excel and display data
                    for i in range(7, sheet.max_row+1):
                        sheet.cell(row=i, column=4).value = sheet.cell(row=i, column=5).value[0:6]
                        sheet.cell(row=i, column=6).value = sheet.cell(row=i, column=7).value.year
                    print("Se aplican formulas para columnas agregadas.")
                    
                wb.save(filename)
                print("Se guardar excel modificado.")
                
                new_file_name = marca + "_" + fecha_archivo + '.xlsx'
                rename = os.path.join(directorio, new_file_name)
                print("Se obtiene nombre de archivo final: " + rename)
                base = os.path.join(default_directory, 'FORMATO - MODELO DE REPORTE VERITRADE.xlsx')
                shutil.copy(base, rename)
                print("Se copia archivo plantilla: " + base)
                
                copiarExcel(filename, rename, aux)
                print("Se copia contenido de archivo descargado a plantilla copia: ")
                result['filename'] = rename
        except Exception as e: 
            result['error'] = str(e)
    else:
        result['error'] = "Ya existe el usuario conectado."
    driver.get("https://business2.veritradecorp.com/login/logout") 
    time.sleep(5)
    driver.quit()
    return result
def enable_download_headless(browser,download_dir):
    browser.command_executor._commands["send_command"] = ("POST", '/session/$sessionId/chromium/send_command')
    params = {'cmd':'Page.setDownloadBehavior', 'params': {'behavior': 'allow', 'downloadPath': download_dir}}
    browser.execute("send_command", params)
def rename_last_downloaded_file(dummy_dir, destination_dir, new_file_name):
    def get_last_downloaded_file_path(dummy_dir):
        """ Return the last modified -in this case last downloaded- file path.

            This function is going to loop as long as the directory is empty.
        """
        while not os.listdir(dummy_dir):
            time.sleep(1)
        return max([os.path.join(dummy_dir, f) for f in os.listdir(dummy_dir)], key=os.path.getctime)

    while '.part' in get_last_downloaded_file_path(dummy_dir):
        time.sleep(1)
    file = get_last_downloaded_file_path(dummy_dir)
    print(file)
    file_extension = pathlib.Path(file).suffix
    print(file_extension)
    new_file_name = new_file_name + file_extension
    rename = os.path.join(destination_dir, new_file_name)
    print(rename)
    #shutil.move(file, rename)
    return file
@app.route('/MoveMail', methods=['POST'])
def MoveMail():
    json = request.json
    host = json['host']
    user = json['user']
    pwd = json['pwd'] 
    carpeta = json['carpeta']

    with MailBox(host).login(user, pwd, initial_folder="INBOX") as mailbox:
        uids = [i.uid for i in mailbox.fetch(AND(seen=True))]
        for uid in uids:
            try:
                mailbox.move(uid, carpeta)
            except Exception as e:
                print(str(e)) 
    return ""
@app.route('/GetInfoMail', methods=['POST'])
def GetInfoMail():
    json = request.json
    host = json['host']
    user = json['user']
    pwd = json['pwd'] 
    uid = json['uid']
    data = {}
    with MailBox(host).login(user, pwd, initial_folder='INBOX') as mailbox:
    #with MailBox('outlook.office365.com').login('testGesco@efc.com.pe', '@Peru2021@', initial_folder='INBOX') as mailbox:
        # COPY all messages from current folder to folder1, *by one
       for msg in mailbox.fetch(AND(uid=uid)):
            with open('mymessage.eml', 'wb') as f:
                f.write(bytes(msg.obj))
            data["uid"] = msg.uid          # str or None: '123'
            data["subject"] = msg.subject      # str: 'some subject 你 привет'
            data["from_"] = msg.from_        # str: 'Bartölke@ya.ru'
            data["to"] = msg.to           # tuple: ('iam@goo.ru', 'friend@ya.ru', )
            data["cc"] = msg.cc           # tuple: ('cc@mail.ru', )
            data["bcc"] = msg.bcc          # tuple: ('bcc@mail.ru', )
            data["reply_to"] = msg.reply_to     # tuple: ('reply_to@mail.ru', )
            data["date"] = msg.date.strftime("%d/%m/%Y")         # datetime.datetime: 1900-1-1 for unparsed, may be naive or with tzinfo
            data["hour"] = msg.date.strftime("%H:%M:%S")         # datetime.datetime: 1900-1-1 for unparsed, may be naive or with tzinfo
            data["date_str"] = msg.date_str     # str: original date - 'Tue, 03 Jan 2017 22:26:59 +0500'
            data["text"] = msg.text         # str: 'Hello 你 Привет'
            data["html"] = msg.html         # str: '<b>Hello 你 Привет</b>'
            data["flags"] = msg.flags        # tuple: ('\\Seen', '\\Flagged', 'ENCRYPTED')
            data["headers"] = msg.headers      # dict: {'received': ('from 1.m.ru', 'from 2.m.ru'), 'anti-virus': ('Clean',)}
            data["size_rfc822"] = msg.size_rfc822  # int: 20664 bytes - size info from server (*useful with headers_only arg)
            data["size"] = msg.size         # int: 20377 bytes - size of received message
            #data["obj"] = msg.obj              # email.message.Message: original object
            data["from_values"] = msg.from_values      # dict or None: {'email': 'im@ya.ru', 'name': 'Ya', 'full': 'Ya <im@ya.ru>'}
            data["sito_valuesze"] = msg.to_values        # tuple: ({'email': '', 'name': '', 'full': ''},)
            data["cc_values"] = msg.cc_values        # tuple: ({'email': '', 'name': '', 'full': ''},)
            data["bcc_values"] = msg.bcc_values       # tuple: ({'email': '', 'name': '', 'full': ''},)
            data["reply_to_values"] = msg.reply_to_values  # tuple: ({'email': '', 'name': '', 'full': ''},)
    return data
        #    res = mailbox.copy(msg.uid, 'INBOX/Copia')

        # MOVE all messages from current folder to folder2, *in bulk (implicit creation of uid list)
        # mailbox.move(mailbox.fetch(), 'INBOX/folder2')
        #mailbox.move(uid, carpeta)
        # DELETE all messages from current folder, *in bulk (explicit creation of uid list)
        # mailbox.delete([msg.uid for msg in mailbox.fetch()])

        # FLAG unseen messages in current folder as Answered and Flagged, *in bulk.
        # flags = (imap_tools.StandardMessageFlags.ANSWERED, imap_tools.StandardMessageFlags.FLAGGED)
        # mailbox.flag(mailbox.fetch('(UNSEEN)'), flags, True)

        # SEEN: mark all messages sent at 05.03.2007 in current folder as unseen, *in bulk
        # mailbox.seen(mailbox.fetch("SENTON 05-Mar-2007"), False)

@app.route('/SaveEML', methods=['POST'])
def SaveEML():
    json = request.json
    host = json['host']
    user = json['user']
    pwd = json['pwd'] 
    uid = json['uid']
    ruta = json['ruta']
    data = {}
    with MailBox(host).login(user, pwd, initial_folder='INBOX') as mailbox:
    #with MailBox('outlook.office365.com').login('testGesco@efc.com.pe', '@Peru2021@', initial_folder='INBOX') as mailbox: 
       for msg in mailbox.fetch(AND(uid=uid)):
            with open(ruta, 'wb') as f:
                f.write(bytes(msg.obj))
    return json
@app.route('/XmlToJson', methods=['POST'])
def XmlToJson():
    entrada = request.json
    xml = entrada['xml']
    obj = xmltodict.parse(xml)
    return json.dumps(obj)
@app.route('/TransformHTML', methods=['POST'])
def transforma_documento():
    json = request.json
    html = json['html']
    xslt = json['xslt']
    dom = ET.fromstring(html)
    transform = ET.XSLT(ET.fromstring(xslt))
    nuevodom = transform(dom)
    resultado = ET.tostring(nuevodom, pretty_print=True)
    return resultado

def TableFromPDFTabula(pdf):
    rnd = random.random()*100000000000
    nombre_pdf = str(rnd) + ".pdf"
    with open(nombre_pdf, 'wb') as theFile:
         theFile.write(base64.b64decode(pdf))
    tables = df = tabula.read_pdf_with_template(nombre_pdf, "template.json")#camelot.read_pdf(nombre_pdf, flavor='stream', edge_tol=200)#, line_scale=40, shift_text=[''])
    i = 1
    lines = ''
    data = {}
    for t in tables:
        soup = BeautifulSoup(t.to_html(), 'html.parser')
        table_rows = soup.find_all('th')
        for th in table_rows:
            if th:  # safety feature
               th.extract()
        lines += str(soup)
    soup = BeautifulSoup(lines, 'html.parser')
    html = soup.prettify()
    return html
@app.route('/HtmlToImage', methods=['POST'])
def HtmlToImage():
    json = request.json
    html = json['html']
    options = {
        'zoom': 2,
        'quality': 100
    }
    imgkit.from_string(html, 'out.jpg', options)
    return ""

@app.route('/PrettyHtml', methods=['POST'])
def PrettyHtml():
    json = request.json
    html = json['html']
    soup = BeautifulSoup(html, 'html.parser')
    return soup.prettify()

def table_json(html_data, xsd_string):
    s = BeautifulSoup(html_data, 'html.parser')
    table = s.find_all('table')
    print(len(table))
    for x in table:
        valid = xml_validator(str(x), xsd_string)
        if valid is True:
           table_MN = pd.read_html(str(x), header=1)
           return x

def xml_validator(some_xml_string, xsd_string):
    try:
        schema_root = etree.XML(xsd_string)
        schema = etree.XMLSchema(schema_root)
        parser = objectify.makeparser(schema=schema)
        objectify.fromstring(some_xml_string, parser)
        return True
    except XMLSyntaxError:
        #handle exception here
        return False
        pass
def validateJson(jsonData, esquema):
    try:
        validate(instance=jsonData, schema=esquema)
    except jsonschema.exceptions.ValidationError as err:
        return False
    return True
def copiarExcel(filename, filename1):
    #filename ="C:\\Users\\Admin\\Desktop\\trading.xlsx"
    wb1 = openpyxl.load_workbook(filename)
    ws1 = wb1.worksheets[0]
      
    # opening the destination excel file 
    #filename1 ="C:\\Users\\Admin\\Desktop\\test.xlsx"
    wb2 = openpyxl.load_workbook(filename1)
    ws2 = wb2.worksheets[0]
      
    # calculate total number of rows and 
    # columns in source excel file
    mr = ws1.max_row
    mc = ws1.max_column
      
    # copying the cell values from source 
    # excel file to destination excel file
    for i in range (1, mr + 1):
        for j in range (1, mc + 1):
            # reading cell value from source excel file
            c = ws1.cell(row = i + 5, column = j)
      
            # writing the read value to destination excel file
            ws2.cell(row = i, column = j).value = c.value
      
    # saving the destination excel file
    wb2.save(str(filename1))
    
#app.run()
app.run(host='0.0.0.0', port=5000)

# import PyPDF2

# pdfFileObj = open('requerimiento.pdf','rb')

# pdfReader = PyPDF2.PdfFileReader(pdfFileObj)

# print("Cantidad de páginas: ", pdfReader.numPages)
# print("Encriptado: ", pdfReader.isEncrypted)

# info = pdfReader.documentInfo

# print("info: ", info)

# for x in range(0, pdfReader.numPages):
    # pageObj = pdfReader.getPage(x)
    # print(pageObj.extractText())